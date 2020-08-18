# -*- coding: utf-8 -*-
# @Time     :2020/02/28  20:51
# @Author   :qingjia
# @File     UtilityTools
import functools
import os
import sys
import json
import codecs
import base64
import unittest

import xlrd
import re
from random import randint
from openpyxl import load_workbook
import collections
from request_parames_template import request_parames_template
# import yaml

lowercase='abcdefghijklmnopqrstuvwxyz'
uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
digist='0123456789'
charList=digist+lowercase
value=""
class UtilityTools():
    # 定义测试库的范围，默认为Test case
    ROBOT_LIBRARY_SCOPE='TEST SUITE'

    def set_encode_utf8(self):
        import sys
        if sys.version.startswith('2'):
            reload(sys)
            sys.setdefaultencoding('utf8')
            # print "python版本号：{}".format(sys.version)
        else:
            pass

    def unicodeToDict(self,input_str):
        """
        将Unicode类型的json转为字典类型
        :param input_str:  需要被转的数据
        :return:  转换后的字典类型数据
        """
        unicodeToString=input_str.encode('utf-8')  #将Unicode转为utf8
        stringToDict=json.loads(unicodeToString)    #将字符串转为json对象
        result=collections.OrderedDict(stringToDict)    #将接送对象转为字典
        return  result


    def get_row_values(self,file_path,sheet_name,row_numbe,star_column=1,null_flag='empty'):
        """
        获取excel某行的值，默认拿所有列的值，返回集合
        :param file_path:
        :param sheet_name:
        :param row_numbe:指定行，从1开始
        :param star_column:指定从那列开始读
        :return:
        """
        # 设置字符编码Python2时使用
        # self.set_encode_utf8()
        # 打开excel
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        # 拿到所有有效列数据行数
        cell_len = sheet.max_column
        # 存储列数据到集合
        column_data_list = []
        for i in range(star_column, cell_len + 1):
            column_data = str(sheet.cell(row_numbe, i).value).strip()
            if str(null_flag)==column_data:
                column_data=''
            column_data_list.append(column_data)
        for data in column_data_list:
            if not 'None'==data:
                return column_data_list
        return None



    def check_filename_sheetname_is_exist(self,file_path,sheet_name):
        '''
        判断文件和sheetname存在
        :param file_path:
        :param sheet_name:
        :return:
        '''
        if not file_path:
            raise RuntimeError("File path is empty")
        if not sheet_name:
            raise RuntimeError(" Sheet name is empty")

    def subInBase(self,base,*subs):
        """
        判断base字符串中包含subs里面的任意一个，存在就返回True，否则返回False\n
        :param base: 主字符串 \n
        :param subs: 被查找的字符串\n
        :return: 返回True或False\n
        """
        self.set_encode_utf8()
        for sub in subs:
            if type(sub)==unicode:
                sub=sub.decode("utf-8")
            sub_str=str(sub).replace(" ","")
            if type(base)==unicode:
                base=base.decode("utf-8")
            if type(base)==dict:
                base=json.dumps(base,ensure_ascii=False).decode("utf-8")
            base_str=str(base).replace(" ","")
            print "处理后的sub_str{}".format(sub_str)
            print "处理后的base_str{}".format(base_str)
            flag=False
            if sub_str in base_str:
                flag=True
                break
        return flag


    def getCellValue(self,file_path,sheet_name, row, column):
        """
        获取某个单元格方法
        :param file_path:文件路径\n
        :param sheet_name:sheet页名称\n
        :param row: 行索引
        :param column: 列索引
        :return:
        """
        # 打开excel
        if not file_path:
            return "File path is empty"
        if not sheet_name:
            return " Sheet name is empty"
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        cellvalue = sheet.cell(row=row, column=column).value
        return cellvalue

    def get_column_values(self,file_path,sheet_name,column_numbe,start_row=1):
        """
        获取excel某列值，默认返回该列所有的值，可以指定，返回集合\n
        :param file_path:文件路径\n
        :param sheet_name:sheet页名称\n
        :param column_numbe: 指定列，从1开始\n
        :param start_row:指定从哪行开始读
        :return:
        """
        # 设置字符编码Python2时使用
        # self.set_encode_utf8()
        # 打开excel
        if not file_path:
            return "File path is empty"
        if not sheet_name:
            return " Sheet name is empty"
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        # 拿到所有有效行数据行数
        rows_len = sheet.max_row
        # 存储列数据到集合
        rows_data_list = []
        for i in range(start_row, rows_len + 1):
            rows_data = sheet.cell(i,column_numbe).value
            rows_data_list.append(rows_data)
        return rows_data_list

    def get_excel_casedata_rownumber(self, file_path, sheet_name, case_id,flag_column=1):
        """
        通过用例id获取excel中指数据的行号，将查询到数据的行号返回
        :param file_path: 文件名\n
        :param sheet_name: sheet页名称\n
        :param case_id: 用例id\n
        :param flag_column: 唯一标识列是第几列\n
        :return: 返回目标数据的行号\n
        """

        # 固定获取列数据方法的文件路径和sheet页名称
        get_column_values = functools.partial(self.get_column_values, file_path, sheet_name)
        # 拿到唯一标识列所有数据
        flag_column_datas = get_column_values(flag_column)
        # print('first_column_datas:{}'.format(flag_column_datas))
        # 拿到指定行的数据
        flag_column_datas_len = len(flag_column_datas)
        target_data_rownumber=None #这里在Python2中必须要先定义下，不然for里面的这个变量会报错
        for i in range(flag_column_datas_len):
            if case_id == flag_column_datas[i]:
                target_data_rownumber=i + 1
                return target_data_rownumber
        return  '{}:case_id not found'.format(case_id)

    def get_excel_casedata_by_caseid(self, file_path, sheet_name, caseid,header_sheetname,assert_sheetname,
                                     headerid_column=3,assertid_column=4,uri_column=5,start_column=6,
                                     title_row_num=1, null_flag='empty'):
        """
        通过用例id获取excel中指定行的数据，将字段名和值组成json格式返回，并将有值的字段组成json参数传入\n
        :param file_path: 文件名\n
        :param sheet_name: sheet页名称\n
        :param caseid: 用例id\n
        :param header_sheetname: 请求头数据存放的sheet页名称\n
        :param headerid_column 请求头id存放的位置\n
        :param assert_sheetname 断言数据存放的sheet页名称\n
        :param assertid_column  断言id存放的位置\n
        :param uri_column  请求路径存放的位置\n
        :param start_column:  数据从那列开始\n
        :param title_row_num: 第几行是title行\n
        :param flag_column: 唯一标识列是第几列\n
        :param null_flag :数据传空标识\n
        :return: 请求路径uri，请求参数，请求头数据，断言数据，json类型数据\n
        """
        self.set_encode_utf8()
        # 固定获取行数据方法的文件路径和sheet页名称
        get_row_values = functools.partial(self.get_row_values, file_path, sheet_name)
        # 拿到标题行数据
        title_row_datas = get_row_values(title_row_num, start_column)
        # 去除title里面的中文备注和空格
        title_row_datas=[re.sub(r'\(.*?\)','',item).strip() for item in title_row_datas]
        # print('title_row_datas:{}'.format(title_row_datas))
        # 拿到数据行数据
        row_number=self.get_excel_casedata_rownumber(file_path,sheet_name,caseid)
        datas_by_rowmunber = get_row_values(row_number, start_column)
        # 处理读取到的用例数据
        title_row_datas_new,targer_data_new=self.del_or_toempty_value(title_row_datas,datas_by_rowmunber,null_flag=null_flag)
        #将标题与请求数据组装成json
        request_pramas=json.dumps(dict([ list(item) for item in zip(title_row_datas_new,targer_data_new)]))
        # 获取到请求的uri
        uri=self.getCellValue(file_path,sheet_name,row_number,uri_column)
        # 获取请求头信息
        header_id=self.getCellValue(file_path, sheet_name, row_number, headerid_column)
        header_data = self.get_excel_assertdata_by_assertid(file_path, header_sheetname, header_id)
        # 获取断言数据
        assertid = self.getCellValue(file_path, sheet_name, row_number, assertid_column)
        assert_data=self.get_excel_assertdata_by_assertid(file_path,assert_sheetname,assertid)
        return  uri,request_pramas,header_data,assert_data


    def get_excel_casedata_all(self, file_path, sheet_name, header_sheetname='headers',assert_sheetname='assertion',
                                     headerid_column=3,assertid_column=4,uri_column=5,method_column=6,start_column=7,
                                     title_row_num=1, null_flag='empty'):
        """
        通过用例id获取excel中指定行的数据，将字段名和值组成json格式返回，并将有值的字段组成json参数传入\n
        :param file_path: 文件名\n
        :param sheet_name: sheet页名称\n
        :param header_sheetname: 请求头数据存放的sheet页名称\n
        :param headerid_column 请求头id存放的位置\n
        :param assert_sheetname 断言数据存放的sheet页名称\n
        :param assertid_column  断言id存放的位置\n
        :param uri_column  请求路径存放的位置\n
        :param method_column  请求方法存放的位置\n
        :param start_column:  数据从那列开始\n
        :param title_row_num: 第几行是title行\n
        :param flag_column: 唯一标识列是第几列\n
        :param null_flag :数据传空标识\n
        :return: 返回所有的测试数据
        """
        self.set_encode_utf8()
        # 固定获取行数据方法的文件路径和sheet页名称
        get_row_values = functools.partial(self.get_row_values, file_path, sheet_name)
        # 拿到标题行数据
        title_row_datas = get_row_values(title_row_num, start_column)
        # 去除title里面的中文备注和空格
        title_row_datas=[re.sub(r'\(.*?\)','',item).strip() for item in title_row_datas]
        # print('title_row_datas:{}'.format(title_row_datas))
        # 拿到所有数据
        wb=load_workbook(file_path)
        sheet=wb[sheet_name]
        max_row_num=sheet.max_row
        testcases=[]
        for row_number in range(2,max_row_num+1):
            case={}
            # 拿到测试数据
            datas_by_rowmunber = get_row_values(row_number, start_column,null_flag)
            request_parames=''
            if datas_by_rowmunber:
                # 处理读取到的用例数据
                # title_row_datas_new,targer_data_new=self.del_or_toempty_value(title_row_datas,datas_by_rowmunber,null_flag=null_flag)
                variable_name_temp=['${%s}' %name for name in title_row_datas]
                name_value_dict=dict(zip(variable_name_temp,datas_by_rowmunber))
                #注意request_parames_template中的属性要和sheetname对应
                request_parames=self.replace_data(request_parames_template.__dict__.get(sheet_name),name_value_dict)
                request_parames=re.sub(r'\s','',request_parames)
            # 获取到请求的uri
            uri=self.getCellValue(file_path,sheet_name,row_number,uri_column)
            # 获取到请求方式
            method=self.getCellValue(file_path,sheet_name,row_number,method_column)
            # 获取请求头信息
            header_id=self.getCellValue(file_path, sheet_name, row_number, headerid_column)
            header_data = self.get_excel_assertdata_by_assertid(file_path, header_sheetname, header_id)
            # 获取断言数据
            assertid = self.getCellValue(file_path, sheet_name, row_number, assertid_column)
            assert_data=self.get_excel_assertdata_by_assertid(file_path,assert_sheetname,assertid)
            # 获取到用例id
            caseid = self.getCellValue(file_path, sheet_name, row_number, 1)
            case['caseid']=caseid
            case['request_parames']=request_parames
            case['uri']=uri
            case['method']=method
            case['header_data']=header_data
            case['assert_data']=assert_data
            testcases.append(case)
        return  testcases


    def get_excel_assertdata_by_assertid(self, file_path, sheet_name, assert_id,start_column=3, title_row_num=1,flag_column=1, null_flag='empty'):
        """
        通过用例id获取该用例的断言数据，将字段名和值组成json格式返回，并将有值的字段组成json参数传入\n
        :param file_path: 文件名\n
        :param sheet_name: sheet页名称\n
        :param assert_id: 断言信息id\n
        :param start_column:  数据从那列开始\n
        :param title_row_num: 第几行是title行\n
        :param flag_column: 唯一标识列是第几列\n
        :param null_flag :数据传空标识\n
        :return: 返回获取到的断言信息json类型数据\n
        """
        # 固定获取行数据方法的文件路径和sheet页名称
        get_row_values = functools.partial(self.get_row_values, file_path, sheet_name)
        # 拿到标题行数据
        title_row_datas = get_row_values(title_row_num, start_column)
        # 去除title里面的中文备注和空格
        title_row_datas=[re.sub(r'\(.*?\)','',item).strip() for item in title_row_datas]
        # print('title_row_datas:{}'.format(title_row_datas))
        #  获取到断言数据
        assert_rownumber=self.get_excel_casedata_rownumber(file_path,sheet_name,assert_id,flag_column)
        datas_by_rowmunber=get_row_values(assert_rownumber, start_column)
        # 处理读取到的用例数据
        title_row_datas_new,targer_data_new=self.del_or_toempty_value(title_row_datas,datas_by_rowmunber,null_flag=null_flag)
        #将标题与请求数据组装成json
        assert_data=json.dumps(dict([ list(item) for item in zip(title_row_datas_new,targer_data_new)]))
        return  assert_data



    def replace_data(self,parameters,variables,expression='\$\{.*?\}'):
        """
        将变量替换为真实数据
        :param parameters: 需要被替换的参数
        :param variable:变量存放数组格式
        :return: 返回替换后的数据
        """
        self.set_encode_utf8()
        # 判断变量中是否有键和值相同的数据,如果出现会造成死循环
        for varialbe_key in variables:
            if varialbe_key==variables[varialbe_key]:
               raise RuntimeError("[{}:{}] error:Same key and value ".format(varialbe_key,variables[varialbe_key]))

        re_data_list = re.findall(expression, parameters)
        if re_data_list:
            for re_data in re_data_list:
                replace_value =str(variables.get(re_data))
                parameters = parameters.replace(re_data, str(replace_value))
        return parameters


    def delete_invalid_data(self,text_data,delete_flag='None'):
        self.set_encode_utf8()
        text_data=str(text_data)
        # 删除参数中没有传的参数字段
        text_data=re.sub(r",(\s*?)'([0-9a-zA-Z_]*?)'(\s*?):(\s*?)[']{0,1}\${\w*?}[']{0,1}|,(\s*?)'([0-9a-zA-Z_]*?)'(\s*?):(\s*?)[']{0,1}"+delete_flag+r"[']{0,1}","",text_data)
        text_data= re.sub(r"'([0-9a-zA-Z_]*?)'(\s*?):(\s*?)[']{0,1}\${\w*?}[']{0,1}(\s*?),|'([0-9a-zA-Z_]*?)'(\s*?):(\s*?)[']{0,1}"+delete_flag+r"[']{0,1}(\s*?),","",text_data)
        text_data=re.sub(r"'([0-9a-zA-Z_]*?)'(\s*?):(\s*?)[']{0,1}\${\w*?}[']{0,1}(\s*?)|'([0-9a-zA-Z_]*?)'(\s*?):(\s*?)[']{0,1}"+delete_flag+r"[']{0,1}(\s*?)","",text_data)

        # 删除参数中没有传的参数字段(双引号的情况)
        text_data=re.sub(r',(\s*?)"([0-9a-zA-Z_]*?)"(\s*?):(\s*?)["]{0,1}\${\w*?}["]{0,1}|,(\s*?)"([0-9a-zA-Z_]*?)"(\s*?):(\s*?)["]{0,1}'+delete_flag+r'["]{0,1}',"",text_data)
        text_data= re.sub(r'"([0-9a-zA-Z_]*?)"(\s*?):(\s*?)["]{0,1}\${\w*?}["]{0,1}(\s*?),|"([0-9a-zA-Z_]*?)"(\s*?):(\s*?)["]{0,1}'+delete_flag+r'["]{0,1}(\s*?),',"",text_data)
        text_data=re.sub(r'"([0-9a-zA-Z_]*?)"(\s*?):(\s*?)["]{0,1}\${\w*?}["]{0,1}(\s*?)|"([0-9a-zA-Z_]*?)"(\s*?):(\s*?)["]{0,1}'+delete_flag+r'["]{0,1}(\s*?)',"",text_data)

        # 删除url中的参数
        text_data=re.sub(r"&(\s*?)([^&]*?)(\s*?)=(\s*?)\${\w*?}|&(\s*?)([^&]*?)(\s*?)=(\s*?)"+delete_flag,"",text_data)
        text_data=re.sub(r"(\s*?)([^&]*?)(\s*?)=(\s*?)\${\w*?}&|(\s*?)([^&]*?)(\s*?)=(\s*?)"+delete_flag+r"(\s*?)&","",text_data)
        text_data=re.sub(r"\?([^&]*?)(\s*?)=(\s*?)\${\w*?}|\?([^&]*?)(\s*?)=(\s*?)"+delete_flag,"?",text_data)
        text_data=re.sub(r"([^&]*?)(\s*?)=(\s*?)\${\w*?}|([^&]*?)(\s*?)=(\s*?)"+delete_flag,"",text_data)

        text_data=re.sub(r"\${\w*?}|"+delete_flag,"",text_data)
        return text_data
    def replace_variable(self,parameters,variables):
        #替换参数中的变量
        parameters=self.replace_data(parameters,variables)
        #替换读取的true
        if not parameters or not variables:
            raise RuntimeError('被替换参数或变量为空,参数:{}\n变量{}'.format(parameters, variables))
        flag = True
        while flag:
            re_data = re.search('(true)[\,,\}]', parameters)
            if re_data:
                re_data_value = re_data.group(1)
                print re_data_value
                parameters = parameters.replace(re_data_value, 'True')
            else:
                flag = False
        return parameters


    def del_or_toempty_value(self,title_data,target_data,null_flag='empty'):
        """
        处理用例数据中不需要传的参数，和需要替换成空的参数
        :param title_data: 参数的key集合
        :param target_data: 参数的值集合
        :param null_flag: 值为空的标记
        :return:
        """
        if  not isinstance(title_data,list):
            return "title_data is not list"
        if  not isinstance(target_data,list):
            return "target_data is not list"
        new_title_data=[]
        new_target_data=[]
        targer_data_len = len(target_data)
        # 将数据中空标记关键子替换成空，None的参数删除
        for i in range(targer_data_len):
            if str(null_flag)==str(target_data[i]).strip():
                    target_data[i]=''
            if target_data[i] is None:
                continue
            else:
                new_title_data.append(title_data[i])
                new_target_data.append(target_data[i])
        return new_title_data,new_target_data


    def ReadTestDate(self,filePath,sheetName,key_colx=0,value_colx=1):
        """
        读取Excel中指定列的数据，默认是第一列和第二列数据\n
        :param filePath:  文件路径\n
        :param sheetName: sheet页名称\n
        :param key_colx: 列索引\n
        :param value_colx: 列索引\n
        :return:返回二维数组[['key', 'value'], ['k1', 'v1'], ['k2', 'v2']]
        """
        #python2中要加
        self.set_encode_utf8()
        if not filePath:
            return 'filePath is empty, please check'

        if not sheetName:
            return 'sheetName is empty, please check'
        else:
            with xlrd.open_workbook(filePath) as wb:
                # sheetName=str(sheetName).decode('utf-8') #Python2中的写法
                sheetName=str(sheetName)
                sheet=wb.sheet_by_name(sheetName)
                keyList=sheet.col_values(colx=key_colx)
                valueList=sheet.col_values(colx=value_colx)
                if len(keyList)!=len(valueList):
                    return 'keys and values are not equal'
                # Python2中要加.decode('utf-8')
                reslist=[[keyList[i],str(valueList[i]).decode('utf-8') if not isinstance(valueList[i],(int,float)) else str(int(valueList[i]))] for i in range(len(keyList))]
                # reslist=[[keyList[i],str(valueList[i]) if not isinstance(valueList[i],(int,float)) else str(int(valueList[i]))] for i in range(len(keyList))]
                return reslist


    def getNeedChangeStrList(self,needChangeStr,splitStr):
        """
        将返回参数里面的kye=value格式的数据拆分成[['k1', 'v1'], ['k12', 'v2']]\n
        :param needChangeStr: 需要被处理的数据\n
        :param splitStr:键值对之前的分割符\n
        :return:返回二维数组
        """
        if needChangeStr:
            # keyvalue.index("=")代表找到字符中=号的位置取=号的索引
            return [[keyvalue[0:keyvalue.index("=")],keyvalue[keyvalue.index('=')+1:]]for keyvalue in str(needChangeStr).split(splitStr)]


    def copyExcel(self,wb):
        from xlutils.filter import process,XLRDReader,XLWTWriter
        w=XLWTWriter()
        process(XLRDReader(wb,'unknown.xls'),w)
        return w.output[0][1],w.style_list


    def getFieldlocation(self,sheetObj,fieldName):
        """
        定位传入fieldName字段所在的行，sheet也中每行必须要有字段标题\n
        :param sheetObj:  sheet页面对象\n
        :param fieldName: 需要寻找的字段\n
        :return: 返回字段所在行的对象\n
        """
        for rowx in range(sheetObj.nrows): #获取sheet页总共的行数并循环
            a=sheetObj.cell_value(rowx,0)
            if a==fieldName: #判断传入的字段与当前行的第一列的值是否相等
                return rowx
        return -1


    def writeTestData(self,filePath,sheetName,needChangeStr,splitStr=':'):
        """
        将处理过的数据写如到excel中\n
        :param filePath: 文件路径\n
        :param sheetName: sheet页名称\n
        :param needChangeStr: 需要处理的数据\n
        :param splitStr: 分割符号\n
        :return:
        """
        sheetIndex=0
        newStrList=self.getNeedChangeStrList(needChangeStr,splitStr)
        # 它提供的formatting_info参数取值为True时（为了节省内存，该参数默认为False），就会读取各种格式的信息。
        rb=xlrd.open_workbook(filename=filePath,formatting_info=True)
        # enumerate用例循环可迭代对象，获取到迭代对象的值和索引
        for index,name in enumerate(rb.sheet_names()):
            sheetName=str(sheetName)
            if name==sheetName:  #找到输入sheetname的索引
                sheetIndex=index
                break
        oldSheet=rb.sheet_by_index(sheetx=sheetIndex)
        wb,styleList=self.copyExcel(rb)

        newSheet=wb.get_sheet(sheetIndex)
        style=styleList[oldSheet.cell_xf_index(0,0)]
        for i in range(len(newStrList)):
            # 返回当前二维数组中的第i个数组中的第一个值在sheet页中的行索引
            fieldRowNum=self.getFieldlocation(oldSheet,newStrList[i][0])
            # 拿到二维数组中第i个数组中的第一个值
            fieldName=newStrList[i][0]
            # 拿到二维数组中第i个数组中的第一个值
            valueList=newStrList[i][1].split(',')
            valueLength=len(valueList)
            if fieldRowNum>=0: #如果行号大于等于0进入
                # 将获取到的第一个值写到当前行的第一列中
                newSheet.write(fieldRowNum,0,fieldName,style)
                if valueLength>1: #如果获取到的值不止一个
                    for j in range(valueLength):
                        # 循环获取到的值写到当前行内
                        newSheet.write(fieldRowNum,1+j,valueList[j],style)
                else:
                    newSheet.write(fieldRowNum,1,valueList,style) #如果没有获取到数据就将数据写到第一行
            else:
                newFieldRow=oldSheet.nrows #获取到旧的sheet页的总行数
                newSheet.write(newFieldRow,0,fieldName,style)
                if valueLength>1:
                    for j in range(valueLength):
                        newSheet.write(newFieldRow,1+j,valueList[j],style)
        rb.release_resources()
        wb.save(filePath)


    def FindValueByJsonPath(self,findBody,findPath,key='all'):
        """
        获取Json中对于路径的值及个数\n
        :param findBody: 被查找的json\n
        :param findPath: 查找路径$.key[].key2\n
        :param key: 返回获取到的第几个值\n
        :return:
        """
        self.set_encode_utf8()
        import jsonpath
        findCount=0
        error=['notFound']
        findReturnStrKey=key
        if not isinstance(findBody,dict):
            json_data=json.loads(findBody)
        else:
            json_data=findBody
        findReturnStr=jsonpath.jsonpath(json_data,findPath)
        if(findReturnStr):
            findCount=str(len(findReturnStr))
            findReturnStrLen=len(findReturnStr)
            if str(findReturnStrKey).upper() == 'ALL':
                return findReturnStr,findCount
            elif int(findReturnStrKey)>=findReturnStrLen:
                findReturnStrKey=findReturnStrLen-1
                return findReturnStr[findReturnStrKey],findCount
            else:
                return findReturnStr[int(findReturnStrKey)],findCount
        return error,findCount


    def regular_match_params(self,expression,datas):
        """
        根据正则表达式获取参数中需要被替换的数据\n
        :param expression: 匹配表达式\n
        :param datas: 被匹配的数据\n
        :return: 匹配后的数据\n
        """
        import re
        global value
        global match_value
        datas=datas.replace(" ","")
        # 判断表达式是否以逗号结尾
        if expression.endswith(','):
            data=re.search(expression,datas) #用有逗号的表达式进行匹配
            expression_temp=expression[0:-1] #去掉表达式的逗号
            if '*?' in expression_temp: #去掉表达式里面的？
                expression_temp=expression_temp.replace("?","")
            data_n=re.search(expression_temp,datas) #用去掉逗号及？的表达式进行匹配
            if data and data_n: #如果有逗号的和没有逗号的表达式都匹配到了就返回有逗号的匹配结果
                value=data.group()
                if "}" in value: #如果匹配的值中包含}就截取}之前的数据
                    index=value.find('}')
                    value=value[0:index]
                # return value
                match_value=value
            if not data and data_n: #如果只有没有逗号的表达式匹配到了就返回没有逗号的匹配结果
                value=data_n.group()
                if "}" in value:# 如果匹配的值中包含}就截取}之前的数据
                    index = value.find('}')
                    value = value[0:index]
                # return value
                match_value = value
            if not data and not data_n:
                # return "No data matched"
                match_value = "No data matched"
        else:
            if '*?' in expression:  # 去掉表达式里面的？
                expression= expression.replace("?", "")
            data_n=re.search(expression,datas)
            expression_temp=expression+','
            data=re.search(expression_temp,datas)
            if data and data_n:  # 如果有逗号的和没有逗号的表达式都匹配到了就返回有逗号的匹配结果
                value = data.group()
                if "}" in value:  # 如果匹配的值中包含}就截取}之前的数据
                    index = value.find('}')
                    value =value[0:index]
                # return value
                match_value = value
            if not data and data_n:  # 如果只有没有逗号的表达式匹配到了就返回没有逗号的匹配结果
                value = data_n.group()
                if "}" in value:  # 如果匹配的值中包含}就截取}之前的数据
                    index = value.find('}')
                    value = value[0:index]
                # return value
                match_value = value
            if not data and not data_n:
                match_value ="No data matched"
        if not match_value.endswith(','):
            match_value=',{}'.format(match_value)
        return match_value


    def delete_params(self,data_json,*keys):
        """
        完成参数中单个或多个字段的删除
        :param keys: 需要删除字段的值，如果是多个必须以list方式传入\n
        :param data_json: 被删除的参数\n
        :return: 返回删除后的参数
        """
        # 替换过程
        self.set_encode_utf8()
        context_log=[]
        data_json=data_json.replace(" ","")
        for i in range(len(keys)):
            expression = '"'+ keys[i] +'.*?,'  # 组装获取需要替换参数的正则表达式
            value = self.regular_match_params(expression, data_json)
            newValue = data_json.replace(value, "")
            log='===================开始删除第【{}】个键值对===================\n原始数据为【{}】\n表达式【{}】\n' \
                '匹配到的数据为【{}】\n替换后的数据为【{}】'.format(i+1,data_json, expression, value,newValue)
            context_log.append(log.decode('utf-8'))
            data_json = newValue
            if ',}' in newValue:
                newValue = newValue.replace(',}', "")
        return newValue,context_log


    # def yaml_to_json(self,filePath):
    #     """
    #     将yaml格式转为json格式
    #     :param filePath: 存放yaml的文件
    #     :return:
    #     """
    #     data=open(filePath)
    #     # 将yaml加载为字典
    #     dict_data=yaml.load(data)
    #     # 将字典转为json
    #     json_data=json.dumps(dict_data)
    #     return json_data
    #
    # def json_to_yaml(self,filePath):
    #     """
    #     将json格式转为yaml
    #     :param filePath:
    #     :return:
    #     """
    #     read_file=open(filePath)
    #     dict_data=yaml.load(read_file)
    #     # target_file=open("yaml_file_path","w")  #定义转换后输出的文件路径
    #     # steam等于None代表输出到控制台，要想指定输出位置可以将steam=target_file
    #     yaml_data=yaml.safe_dump(dict_data,stream=None,default_flow_style=False)
    #     return  yaml_data

if __name__ == '__main__':
    tools=UtilityTools()
    # # print tools.ReadTestDate('test01.xlsx','Sheet1')
    # """
    # 1.写一个从excel里面选择数据的方法，存到一个二维数组
    # 2.编写替换数据的方法
    # 3.编写清空未填写值的参数
    # 4.可以设置空值的变量
    # """
    # # 打开excel
    #
    # file_path = r'../03TestData/testdata_sit.xlsx'
    # sheet_name = 'login'
    # flag_id = 'login_case1'
    # header_sheetname='headers'
    # assert_sheetname='assertion'
    # # uri,param=tools.get_excel_casedata_by_caseid(file_path, sheet_name, flag_id)
    # # row_num=tools.get_excel_casedata_rownumber(file_path,sheet_name,flag_id)
    # # print row_num
    # # uri,request_pramas,header,assert_data=tools.get_excel_casedata_by_caseid(file_path,sheet_name,flag_id,header_sheetname,assert_sheetname)
    # # print uri
    # # print request_pramas
    # # print header
    # # print assert_data
    # # varialbe1={u'${version}': u'1.1', u'${usertoken}': u'token111', u'${orderno}': u'1234567'}
    # varialbe={'${real_pwd}': '111', '${real_number}':'1234567'}
    # # print("param:{}".format(param))
    # # var=tools.replace_variable(request_pramas,varialbe)
    # # print(var)
    # # expression='"mobilephone.*?[",\]]'
    # # data = re.search(expression, parameters)
    # # print data.group()
    # # # print data
    # parameters='{"msg": "登录成功", "code": 10001, "http_code": 200}'
    # parameters,context_log=tools.delete_params(parameters,'http_code')
    # a=',"http_code": 200'

    file_path=r'../03TestData/testdata_sit.xlsx'
    sheet_name='createVolume'
    data=tools.get_excel_casedata_all(file_path=file_path,sheet_name=sheet_name)
    print data
    data=tools.delete_invalid_data(data)
    print data



