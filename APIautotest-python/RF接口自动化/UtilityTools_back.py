# -*- coding: utf-8 -*-
# @Time     :2020/02/28  20:51
# @Author   :qingjia
# @File     UtilityTools
import functools
import os
import sys
import json
import codecs
import base64
import unittest

import xlrd
import re
from random import randint
from openpyxl import load_workbook
import collections
# import yaml

lowercase='abcdefghijklmnopqrstuvwxyz'
uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
digist='0123456789'
charList=digist+lowercase
value=""
class UtilityTools():
    # 定义测试库的范围，默认为Test case
    ROBOT_LIBRARY_SCOPE='TEST SUITE'

    def set_encode_utf8(self):
        import sys
        reload(sys)
        sys.setdefaultencoding('utf8')

    def unicodeToDict(self,input_str):
        """
        将Unicode类型的json转为字典类型
        :param input_str:  需要被转的数据
        :return:  转换后的字典类型数据
        """
        unicodeToString=input_str.encode('utf-8')  #将Unicode转为utf8
        stringToDict=json.loads(unicodeToString)    #将字符串转为json对象
        result=collections.OrderedDict(stringToDict)    #将接送对象转为字典
        return  result


    def get_row_values(self,file_path,sheet_name,row_numbe,star_column=1):
        """
        获取excel某行的值，默认拿所有列的值，返回集合
        :param file_path:
        :param sheet_name:
        :param row_numbe:指定行，从1开始
        :param star_column:指定从那列开始读
        :return:
        """
        # 设置字符编码Python2时使用
        # self.set_encode_utf8()
        # 打开excel
        if not file_path:
            return "File path is empty"
        if not sheet_name:
            return  " Sheet name is empty"
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        # 拿到所有有效列数据行数
        cell_len = sheet.max_column
        # 存储列数据到集合
        column_data_list = []
        for i in range(star_column, cell_len + 1):
            column_data = sheet.cell(row_numbe, i).value
            # if column_data is None:
            #     column_data=''
            column_data_list.append(column_data)
        return column_data_list

    def getCellValue(self,file_path,sheet_name, row, column):
        """
        获取某个单元格方法
        :param file_path:文件路径\n
        :param sheet_name:sheet页名称\n
        :param row: 行索引
        :param column: 列索引
        :return:
        """
        # 打开excel
        if not file_path:
            return "File path is empty"
        if not sheet_name:
            return " Sheet name is empty"
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        cellvalue = sheet.cell(row=row, column=column).value
        return cellvalue

    def get_column_values(self,file_path,sheet_name,column_numbe,start_row=1):
        """
        获取excel某列值，默认返回该列所有的值，可以指定，返回集合\n
        :param file_path:文件路径\n
        :param sheet_name:sheet页名称\n
        :param column_numbe: 指定列，从1开始\n
        :param start_row:指定从哪行开始读
        :return:
        """
        # 设置字符编码Python2时使用
        # self.set_encode_utf8()
        # 打开excel
        if not file_path:
            return "File path is empty"
        if not sheet_name:
            return " Sheet name is empty"
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        # 拿到所有有效行数据行数
        rows_len = sheet.max_row
        # 存储列数据到集合
        rows_data_list = []
        for i in range(start_row, rows_len + 1):
            rows_data = sheet.cell(i,column_numbe).value
            rows_data_list.append(rows_data)
        return rows_data_list

    def get_excel_casedata_by_caseid(self, file_path, sheet_name, case_id, uri_column=3,start_column=4, title_row_num=1, flag_column=1, null_flag='empty'):
        """
        通过用例id获取excel中指定行的数据，将字段名和值组成json格式返回，并将有值的字段组成json参数传入\n
        :param file_path: 文件名\n
        :param sheet_name: sheet页名称\n
        :param case_id: 用例id\n
        :param start_column:  数据从那列开始\n
        :param title_row_num: 第几行是title行\n
        :param flag_column: 唯一标识列是第几列\n
        :param null_flag :数据传空标识\n
        :return: 返回json类型数据\n
        """
        # 固定获取行数据方法的文件路径和sheet页名称
        get_row_values = functools.partial(self.get_row_values, file_path, sheet_name)
        # 固定获取列数据方法的文件路径和sheet页名称
        get_column_values = functools.partial(self.get_column_values, file_path, sheet_name)
        # 拿到标题行数据
        title_row_datas = get_row_values(title_row_num, start_column)
        # title_row_datas=re.sub(r'\(.*?\)','1',str(title_row_datas))
        # 去除title里面的中文备注和空格
        title_row_datas=[re.sub(r'\(.*?\)','',item).strip() for item in title_row_datas]
        # print('first_row_datas:{}'.format(title_row_datas))
        # 拿到唯一标识列所有数据
        flag_column_datas = get_column_values(flag_column)
        # print('first_column_datas:{}'.format(flag_column_datas))
        # 拿到指定行的数据
        flag_column_datas_len = len(flag_column_datas)
        datas_by_caseid=[] #这里在Python2中必须要先定义下，不然for里面的这个变量会报错
        uri_by_caseid=''
        for i in range(flag_column_datas_len):
            if case_id == flag_column_datas[i]:
                datas_by_caseid = get_row_values(i + 1,start_column)
                uri_by_caseid=self.getCellValue(file_path, sheet_name,i + 1,uri_column)
        # 处理读取到的用例数据
        title_row_datas,targer_data_new=self.del_or_toempty_value(title_row_datas,datas_by_caseid,null_flag=null_flag)
        #将标题与用例数据组装成json返回
        return  uri_by_caseid,json.dumps(dict([ list(item) for item in zip(title_row_datas,targer_data_new)]))

    def replace_variable(self,parameters,variables):
        """
        将变量替换为真实数据
        :param parameters: 需要被替换的参数
        :param variable:变量存放数组格式
        :return: 返回替换后的数据
        """
        self.set_encode_utf8()
        # 判断变量中是否有键和值相同的数据,如果出现会造成死循环
        for varialbe_key in variables:
            if varialbe_key==variables[varialbe_key]:
                return "[{}:{}] error:Same key and value ".format(varialbe_key,variables[varialbe_key])
        flag = True
        while flag:
            re_data = re.search('\$\{(.*?)\}', parameters)
            if re_data:
                re_data_value = re_data.group()
                replace_value=variables.get(re_data_value)
                if not replace_value:
                    return "{} variable not defined ".format(re_data_value)
                parameters = parameters.replace(re_data_value,replace_value)
            else:
                flag = False
        return parameters

    def del_or_toempty_value(self,title_data,target_data,null_flag='empty'):
        """
        处理用例数据中不需要传的参数，和需要替换成空的参数
        :param title_data: 参数的key集合
        :param target_data: 参数的值集合
        :param null_flag: 值为空的标记
        :return:
        """
        if  not isinstance(title_data,list):
            return "title_data is not list"
        if  not isinstance(target_data,list):
            return "target_data is not list"
        new_title_data=[]
        new_target_data=[]
        targer_data_len = len(target_data)
        # 将数据中空标记关键子替换成空，None的参数删除
        for i in range(targer_data_len):
            if str(null_flag)==str(target_data[i]).strip():
                    target_data[i]=''
            if target_data[i] is None:
                continue
            else:
                new_title_data.append(title_data[i])
                new_target_data.append(target_data[i])
        return new_title_data,new_target_data


    def ReadTestDate(self,filePath,sheetName,key_colx=0,value_colx=1):
        """
        读取Excel中指定列的数据，默认是第一列和第二列数据\n
        :param filePath:  文件路径\n
        :param sheetName: sheet页名称\n
        :param key_colx: 列索引\n
        :param value_colx: 列索引\n
        :return:返回二维数组[['key', 'value'], ['k1', 'v1'], ['k2', 'v2']]
        """
        #python2中要加
        self.set_encode_utf8()
        if not filePath:
            return 'filePath is empty, please check'

        if not sheetName:
            return 'sheetName is empty, please check'
        else:
            with xlrd.open_workbook(filePath) as wb:
                # sheetName=str(sheetName).decode('utf-8') #Python2中的写法
                sheetName=str(sheetName)
                sheet=wb.sheet_by_name(sheetName)
                keyList=sheet.col_values(colx=key_colx)
                valueList=sheet.col_values(colx=value_colx)
                if len(keyList)!=len(valueList):
                    return 'keys and values are not equal'
                # Python2中要加.decode('utf-8')
                # reslist=[[keyList[i],str(valueList[i]).decode('utf-8') if not isinstance(valueList[i],(int,float)) else str(int(valueList[i]))] for i in range(len(keyList))]
                reslist=[[keyList[i],str(valueList[i]) if not isinstance(valueList[i],(int,float)) else str(int(valueList[i]))] for i in range(len(keyList))]
                return reslist


    def getNeedChangeStrList(self,needChangeStr,splitStr):
        """
        将返回参数里面的kye=value格式的数据拆分成[['k1', 'v1'], ['k12', 'v2']]\n
        :param needChangeStr: 需要被处理的数据\n
        :param splitStr:键值对之前的分割符\n
        :return:返回二维数组
        """
        if needChangeStr:
            # keyvalue.index("=")代表找到字符中=号的位置取=号的索引
            return [[keyvalue[0:keyvalue.index("=")],keyvalue[keyvalue.index('=')+1:]]for keyvalue in str(needChangeStr).split(splitStr)]


    def copyExcel(self,wb):
        from xlutils.filter import process,XLRDReader,XLWTWriter
        w=XLWTWriter()
        process(XLRDReader(wb,'unknown.xls'),w)
        return w.output[0][1],w.style_list


    def getFieldlocation(self,sheetObj,fieldName):
        """
        定位传入fieldName字段所在的行，sheet也中每行必须要有字段标题\n
        :param sheetObj:  sheet页面对象\n
        :param fieldName: 需要寻找的字段\n
        :return: 返回字段所在行的对象\n
        """
        for rowx in range(sheetObj.nrows): #获取sheet页总共的行数并循环
            a=sheetObj.cell_value(rowx,0)
            if a==fieldName: #判断传入的字段与当前行的第一列的值是否相等
                return rowx
        return -1


    def writeTestData(self,filePath,sheetName,needChangeStr,splitStr=':'):
        """
        将处理过的数据写如到excel中\n
        :param filePath: 文件路径\n
        :param sheetName: sheet页名称\n
        :param needChangeStr: 需要处理的数据\n
        :param splitStr: 分割符号\n
        :return:
        """
        sheetIndex=0
        newStrList=self.getNeedChangeStrList(needChangeStr,splitStr)
        # 它提供的formatting_info参数取值为True时（为了节省内存，该参数默认为False），就会读取各种格式的信息。
        rb=xlrd.open_workbook(filename=filePath,formatting_info=True)
        # enumerate用例循环可迭代对象，获取到迭代对象的值和索引
        for index,name in enumerate(rb.sheet_names()):
            sheetName=str(sheetName)
            if name==sheetName:  #找到输入sheetname的索引
                sheetIndex=index
                break
        oldSheet=rb.sheet_by_index(sheetx=sheetIndex)
        wb,styleList=self.copyExcel(rb)

        newSheet=wb.get_sheet(sheetIndex)
        style=styleList[oldSheet.cell_xf_index(0,0)]
        for i in range(len(newStrList)):
            # 返回当前二维数组中的第i个数组中的第一个值在sheet页中的行索引
            fieldRowNum=self.getFieldlocation(oldSheet,newStrList[i][0])
            # 拿到二维数组中第i个数组中的第一个值
            fieldName=newStrList[i][0]
            # 拿到二维数组中第i个数组中的第一个值
            valueList=newStrList[i][1].split(',')
            valueLength=len(valueList)
            if fieldRowNum>=0: #如果行号大于等于0进入
                # 将获取到的第一个值写到当前行的第一列中
                newSheet.write(fieldRowNum,0,fieldName,style)
                if valueLength>1: #如果获取到的值不止一个
                    for j in range(valueLength):
                        # 循环获取到的值写到当前行内
                        newSheet.write(fieldRowNum,1+j,valueList[j],style)
                else:
                    newSheet.write(fieldRowNum,1,valueList,style) #如果没有获取到数据就将数据写到第一行
            else:
                newFieldRow=oldSheet.nrows #获取到旧的sheet页的总行数
                newSheet.write(newFieldRow,0,fieldName,style)
                if valueLength>1:
                    for j in range(valueLength):
                        newSheet.write(newFieldRow,1+j,valueList[j],style)
        rb.release_resources()
        wb.save(filePath)


    def FindValueByJsonPath(self,findBody,findPath,key='all'):
        """
        获取Json中对于路径的值及个数\n
        :param findBody: 被查找的json\n
        :param findPath: 查找路径$.key[].key2\n
        :param key: 返回获取到的第几个值\n
        :return:
        """
        import jsonpath
        findCount=0
        error=['notFound']
        findReturnStrKey=key
        if not isinstance(findBody,dict):
            json_data=json.loads(findBody)
        else:
            json_data=findBody
        findReturnStr=jsonpath.jsonpath(json_data,findPath)
        if(findReturnStr):
            findCount=str(len(findReturnStr))
            findReturnStrLen=len(findReturnStr)
            if str(findReturnStrKey).upper() == 'ALL':
                return findReturnStr,findCount
            elif int(findReturnStrKey)>=findReturnStrLen:
                findReturnStrKey=findReturnStrLen-1
                return findReturnStr[findReturnStrKey],findCount
            else:
                return findReturnStr[int(findReturnStrKey)],findCount
        return error,findCount


    def regular_match_params(self,expression,datas):
        """
        根据正则表达式获取参数中需要被替换的数据\n
        :param expression: 匹配表达式\n
        :param datas: 被匹配的数据\n
        :return: 匹配后的数据\n
        """
        import re
        global value
        # 判断表达式是否以逗号结尾
        if expression.endswith(','):
            data=re.search(expression,datas) #用有逗号的表达式进行匹配
            expression_temp=expression[0:-1] #去掉表达式的逗号
            if '*?' in expression_temp: #去掉表达式里面的？
                expression_temp=expression_temp.replace("?","")
            data_n=re.search(expression_temp,datas) #用去掉逗号及？的表达式进行匹配
            if data and data_n: #如果有逗号的和没有逗号的表达式都匹配到了就返回有逗号的匹配结果
                value=data.group()
                if "}" in value: #如果匹配的值中包含}就截取}之前的数据
                    index=value.find('}')
                    value=value[0:index]
                return value
            if not data and data_n: #如果只有没有逗号的表达式匹配到了就返回没有逗号的匹配结果
                value=data_n.group()
                if "}" in value:# 如果匹配的值中包含}就截取}之前的数据
                    index = value.find('}')
                    value = value[0:index]
            return value
            if not data and not data_n:
                return "No data matched"
        else:
            if '*?' in expression:  # 去掉表达式里面的？
                expression= expression.replace("?", "")
            data_n=re.search(expression,datas)
            expression_temp=expression+','
            data=re.search(expression_temp,datas)
            if data and data_n:  # 如果有逗号的和没有逗号的表达式都匹配到了就返回有逗号的匹配结果
                value = data.group()
                if "}" in value:  # 如果匹配的值中包含}就截取}之前的数据
                    index = value.find('}')
                    value = value[0:index]
                return value
            if not data and data_n:  # 如果只有没有逗号的表达式匹配到了就返回没有逗号的匹配结果
                value = data_n.group()
                if "}" in value:  # 如果匹配的值中包含}就截取}之前的数据
                    index = value.find('}')
                    value = value[0:index]
            return value
            if not data and not data_n:
                return "No data matched"


    def replce_params(self,keys,data_json):
        """
        完成参数中单个或多个字段的删除，多个必须以list传入[k1,k2]\n
        :param keys: 需要删除字段的值，如果是多个必须以list方式传入\n
        :param data_json: 被删除的参数\n
        :return: 返回删除后的参数
        """
        # 替换过程
        context_log=[]
        if not isinstance(keys,list):
            # "key.*?,
            expression = '"'+keys +'.*?,' #组装获取需要替换参数的正则表达式
            value=tools.regular_match_params(expression,data_json)
            newValue = data_json.replace(value,"")
            log='*******原始数据为【{}】;\n表达式【{}】，匹配到的数据为【{}】,替换后的数据为【{}】'.format(data_json,expression,value,newValue)
            context_log.append(log)
            return newValue
        for i in range(len(keys)):
            expression = '"'+ keys[i] +'.*?,'  # 组装获取需要替换参数的正则表达式
            value = tools.regular_match_params(expression, data_json)
            newValue = data_json.replace(value, "")
            log='*******开始第【{}】次原始数据为【{}】;表达式【{}】，匹配到的数据为【{}】,替换后的数据为【{}】'.format(i+1,data_json, expression, value,newValue)
            context_log.append(log)
            data_json = newValue
            if ',}' in newValue:
                newValue = newValue.replace(',}', "")
        return newValue,context_log

    # def yaml_to_json(self,filePath):
    #     """
    #     将yaml格式转为json格式
    #     :param filePath: 存放yaml的文件
    #     :return:
    #     """
    #     data=open(filePath)
    #     # 将yaml加载为字典
    #     dict_data=yaml.load(data)
    #     # 将字典转为json
    #     json_data=json.dumps(dict_data)
    #     return json_data
    #
    # def json_to_yaml(self,filePath):
    #     """
    #     将json格式转为yaml
    #     :param filePath:
    #     :return:
    #     """
    #     read_file=open(filePath)
    #     dict_data=yaml.load(read_file)
    #     # target_file=open("yaml_file_path","w")  #定义转换后输出的文件路径
    #     # steam等于None代表输出到控制台，要想指定输出位置可以将steam=target_file
    #     yaml_data=yaml.safe_dump(dict_data,stream=None,default_flow_style=False)
    #     return  yaml_data

if __name__ == '__main__':
    tools=UtilityTools()
    # print tools.ReadTestDate('test01.xlsx','Sheet1')
    """
    1.写一个从excel里面选择数据的方法，存到一个二维数组
    2.编写替换数据的方法
    3.编写清空未填写值的参数
    4.可以设置空值的变量
    """
    # 打开excel

    file_path = r'../03TestData/testdata_sit.xlsx'
    sheet_name = 'login'
    flag_id = 'login_case1'
    uri,param=tools.get_excel_casedata_by_caseid(file_path, sheet_name, flag_id)
    print uri
    print param
    varialbe1={u'${version}': u'1.1', u'${usertoken}': u'token111', u'${orderno}': u'1234567'}
    # varialbe={u'${version}': u'1.1', u'${orderno}': u'1234567'}
    # print("param:{}".format(param))
    var=tools.replace_variable(param,varialbe1)
    print(var)



