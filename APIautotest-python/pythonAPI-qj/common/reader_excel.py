# -*- coding: utf-8 -*-
# @Time     :2019/09/01  20:56
# @Author   :qingjia
# @File     reader_excel.py 读取excel文件数据

import  openpyxl
from common.logger import logger
class Case:
    # 这个类用来存储用例,用例的数据设为对象的属性

    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型-->[(key,value),(key1,value1)....]
        """
        for item in attrs:
            #第一个是传的实例对象，第二个传的是属性名，第三个是属性值
            setattr(self, item[0], item[1])
class ReadExcel(object):
    """读取excel数据"""
    def __init__(self,file_name,sheet_name):
        """初始化类时将文件对象和sheet页对象传入
        :param file_name  文件名-->str
        :param sheet_name  表单名-->str
        """
        self.file_name=file_name
        self.sheet_name=sheet_name
    def open(self):
        # 加载文件
        self.wb=openpyxl.load_workbook(self.file_name)
        #选择表单
        self.sheet=self.wb[self.sheet_name]
    def close(self):
        #关闭工作簿
        self.wb.close()
    #读取excel所有列所有行数据
    def read_data_obj(self):
        #打开文件操作sheet页
        self.open()
        # 按行读取数据,转换成list，每行数据都存在一个元组中
        rows_data=list(self.sheet.rows)
        # print(rows_data)
        #获取表头信息保存到数组中
        titles=[]
        for title in rows_data[0]:#取出第一行数据，因为第一行就是表头
            titles.append(title.value)#拿出第一行每一列的值添加到表头中
        #开始读取数据部分，创建一个数组保存数据
        cases=[]
        for case in rows_data[1:]:#第二行开始到最后一行
            #创建一个数组保存每列的值
            data=[]
            #循环行中的每一列取出值
            for cell in case:
                data.append(cell.value)# 取出当前行中每列的值放在data中
                #将标题做为键将数据做为值打包成一个标题一个值的元祖("key","value")
            case_data=zip(titles,data)
            #将打包后的数据保存到对象中，每一行数据一个对像
            case_obj=Case(case_data)
            cases.append(case_obj)#将对象保存到数组中
        self.close()#关闭工作簿
        return cases #返回保存好的数据对象

    def readcell_data_obj(self,cells=[]):
        """
       按指定的列，读取excel中的数据，以列表的形式返回，列表中每个对象为一条用例
       excel中的表头为对象的属性，对应的数据为属性值
       :param list1: list  -->要读取的列[1,3,5....]
       :return: type:list--->[case_obj1,case_obj2....]，
        """
        #打开文件操作sheet页
        self.open()
        if len(cells)==0:
            return self.read_data_obj() #返回sheet页中所有数据
        if  not isinstance(cells,list):
            logger.error("请将要获取的列号以数组的方式传入，[{}]".format(cells))
        if isinstance(cells,list) and len(cells) >0:
            # if len(cells)!=0:
                #获取到最大行数max_row
                max_r=self.sheet.max_row
                # 定义一个空列表，用来存放所有用例
                cases = []
                # 定义一个空列表，用来存放表头
                titles = []
                # 循环没一行数据
                for row in range(1,max_r+1):
                    if row==1:#取出第一行数据，因为第一行就是表头
                        for cell in cells:#循环传入的指定列
                            titles.append(self.sheet.cell(row,cell).value)#拿出第一行每一列的值添加到表头中
                    else:
                        #开始读取数据部分，创建一个数组保存数据
                        case_data=[]
                        for cell in cells:#循环传入的指定列
                            case_data.append(self.sheet.cell(row,cell).value)#拿出数据行指定列的值添加到case_data中
                        case=zip(titles,case_data)
                    #将打包后的数据保存到对象中，每一行数据一个对像
                        case_obj=Case(case)
                        cases.append(case_obj)#将对象保存到数组中
                self.close()#关闭工作簿
                return cases #返回保存好的数据对象



if __name__ == '__main__':
    path=r"D:\PycharmProjects\pythonAPI\pythonAPI-qj\data\cases.xlsx"
    read=ReadExcel(path,"register")
    # case=read.read_data_obj()
    # print(case[0].url)

    case=read.readcell_data_obj([1,2])
    print(list(case))
    for c in case:
        print(c.interface)