"""
============================
author:MuSen
time:2019/5/24
E-mail:3247119728@qq.com
============================
"""
import openpyxl


class Case:
    # 这个类用来存储用例,用例的数据设为对象的属性

    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型-->[(key,value),(key1,value1)....]
        """
        for item in attrs:
            #设置属性值，第一个为类对象，第二个为属性名，第三个为属性值
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        # 打开工作簿
        self.wb = openpyxl.load_workbook(self.file_name)
        # 选择表单
        self.sheet = self.wb[self.sheet_name]

    def close(self):
        # 关闭工作簿
        self.wb.close()

    def read_data_obj(self):
        self.open()
        """
        按行读取excel中的数据，以列表的形式返回，列表中每个对象为一条用例
        excel中的表头为对象的属性，对应的数据为属性值
        :return: type:list--->[case_obj1,case_obj2....]，
        """
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        print(rows_data)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]: #每一行都保存在一个元组中
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例对象
        cases = []
        # 判断是否是读取所有数据
        for case in rows_data[1:]:
            # data用来临时存放每一行的用例数据
            data = []
            # 判断该单元格是否为字符串类型，
            for cell in case:
                data.append(cell.value)
            # 将该表头和对应的值保存到元组中（title1,value1）
            case_data = zip(titles, data)
            # 将打包后的元祖数据传入到Case类中，每条数据创建一个Case类的对象，
            case_obj = Case(case_data)
            # 将每条数据的对象保存到cases列表中
            cases.append(case_obj)
        # 关闭工作簿
        self.close()
        return cases

    def readline_data_obj(self, list1=None):
        """
        按指定的列，读取excel中的数据，以列表的形式返回，列表中每个对象为一条用例
        excel中的表头为对象的属性，对应的数据为属性值
        :param list1: list  -->要读取的列[1,2....]
        :return: type:list--->[case_obj1,case_obj2....]，
        """
        self.open()
        list1 = eval(list1)
        # 添加容错机制，如果列表长度为0时返回所有的数据
        if len(list1) == 0:
            return self.read_data_obj()
        # 获取都最大行
        max_r = self.sheet.max_row
        # 定义一个空列表，用来存放所有用例
        cases = []
        # 定义一个空列表，用来存放表头
        titles = []
        # 遍历出所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row == 1:
                # 遍历list1指定的列，获取表头
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    # 将数据添加到表头中
                    titles.append(title)
            else:
                # 定义一个空列表，用来存放该行的数据
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                # 将该条数据和表头进行打包组合，
                case = list(zip(titles, case_data))
                # 创建一个用例对象，将表头和数据传进入初始化
                case_obj = Case(case)
                cases.append(case_obj)
        self.close()
        return cases

    def write_data(self, row, column, msg):
        self.open()
        # 写入数据
        self.sheet.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)
        self.close()



if __name__ == '__main__':
    r = ReadExcel('../data/cases.xlsx', 'register')
    data = r.read_data_obj()
    for i in data:
        print(i.case_id)